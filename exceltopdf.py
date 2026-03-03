import argparse
import os
import shutil
import subprocess
import tempfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties


def find_office_binary() -> str | None:
    for name in ("soffice", "libreoffice"):
        path = shutil.which(name)
        if path:
            return path
    return None



# python exceltopdf.py --input "/home/harold-coder/Desktop/compliance/op_no_chronology_sorted_noncompliant.xlsx"



def apply_print_friendly_layout(input_path: Path) -> Path:
    wb = load_workbook(input_path)
    for ws in wb.worksheets:
        if ws.max_row <= 1 and ws.max_column <= 1 and ws["A1"].value in (None, ""):
            continue
        last_col = get_column_letter(max(1, ws.max_column))
        last_row = max(1, ws.max_row)
        ws.print_area = f"A1:{last_col}{last_row}"
        if ws.sheet_properties.pageSetUpPr is None:
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        else:
            ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.sheet_view.zoomScale = 100
        ws.print_options.horizontalCentered = False
        ws.print_options.verticalCentered = False
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.4
        ws.page_margins.bottom = 0.4

    fd, tmp_name = tempfile.mkstemp(prefix="excel_print_", suffix=".xlsx")
    os.close(fd)
    tmp_path = Path(tmp_name)
    wb.save(tmp_path)
    wb.close()
    return tmp_path


def convert_excel_to_pdf(input_path: Path, output_path: Path | None = None) -> Path:
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")
    if input_path.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
        raise ValueError("Input must be an Excel file (.xlsx/.xlsm/.xls).")

    office_bin = find_office_binary()
    if not office_bin:
        raise RuntimeError(
            "LibreOffice/soffice is not installed or not in PATH. "
            "Install LibreOffice and re-run this script."
        )

    if output_path is None:
        output_path = input_path.with_suffix(".pdf")
    output_path = output_path.resolve()
    outdir = output_path.parent
    outdir.mkdir(parents=True, exist_ok=True)

    temp_input = apply_print_friendly_layout(input_path)
    cmd = [
        office_bin,
        "--headless",
        "--invisible",
        "--nodefault",
        "--nolockcheck",
        "--nologo",
        "--norestore",
        "--convert-to",
        "pdf:calc_pdf_Export",
        "--outdir",
        str(outdir),
        str(temp_input.resolve()),
    ]
    runtime_dir = Path("/tmp/soffice-runtime").resolve()
    profile_dir = Path("/tmp/soffice-profile").resolve()
    runtime_dir.mkdir(parents=True, exist_ok=True)
    profile_dir.mkdir(parents=True, exist_ok=True)
    env = dict(**os.environ)
    env["XDG_RUNTIME_DIR"] = str(runtime_dir)
    env["HOME"] = str(profile_dir)
    env["SAL_USE_VCLPLUGIN"] = "svp"
    env.pop("DISPLAY", None)

    run = subprocess.run(cmd, capture_output=True, text=True, env=env)
    if run.returncode != 0:
        raise RuntimeError(
            "Conversion failed.\n"
            f"Command: {' '.join(cmd)}\n"
            f"stdout: {run.stdout}\n"
            f"stderr: {run.stderr}"
        )

    generated = outdir / f"{temp_input.stem}.pdf"
    if not generated.exists():
        if temp_input.exists():
            temp_input.unlink(missing_ok=True)
        raise RuntimeError(
            "Conversion command completed but expected PDF was not found. "
            f"Expected: {generated}\nstdout: {run.stdout}\nstderr: {run.stderr}"
        )

    if generated.resolve() != output_path:
        if output_path.exists():
            output_path.unlink()
        generated.rename(output_path)
    if temp_input.exists():
        temp_input.unlink(missing_ok=True)

    return output_path


def main():
    parser = argparse.ArgumentParser(description="Convert an Excel file to print-ready PDF.")
    parser.add_argument("--input", required=True, help="Path to input .xlsx/.xlsm/.xls file")
    parser.add_argument("--output", default="", help="Optional output PDF path")
    args = parser.parse_args()

    input_path = Path(args.input).expanduser()
    output_path = Path(args.output).expanduser() if args.output else None
    pdf_path = convert_excel_to_pdf(input_path, output_path)
    print(f"Done. PDF created at: {pdf_path}")


if __name__ == "__main__":
    main()
