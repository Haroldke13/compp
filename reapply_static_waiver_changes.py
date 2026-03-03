#!/usr/bin/env python3
import re
from datetime import datetime
from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = BASE_DIR / "static" / "files" / "Waiver summary, 2025.xlsx"
TARGET_SHEETS = {
    "WSapril,25",
    "wsmay25",
    "june, 2025",
    "july, 2025",
    "August, 2025",
    "sept_oct_2025",
    "Nov_Dec_2025",
    "jan,2026",
}


def normalize(text: object) -> str:
    return " ".join(re.sub(r"[^a-z0-9]+", " ", str(text or "").lower()).split())


def parse_amount(value: object) -> float:
    text = str(value or "").strip()
    if not text:
        return 0.0
    compact = text.lower().replace(",", "")
    short = re.search(r"(-?\d+(?:\.\d+)?)\s*([km])\b", compact)
    if short:
        base = float(short.group(1))
        mult = 1000.0 if short.group(2) == "k" else 1_000_000.0
        return max(0.0, base * mult)
    compact = re.sub(r"[^\d.\-]+", "", compact)
    try:
        return max(0.0, float(compact))
    except Exception:
        return 0.0


def format_amount(value: float) -> str:
    if abs(value - round(value)) < 1e-9:
        return str(int(round(value)))
    return f"{value:.2f}"


def extract_comment_amounts(comment_text: str) -> dict[str, object]:
    text = str(comment_text or "")
    norm = normalize(text)
    out = {
        "rejected": any(t in norm for t in ("reject", "rejected", "declined", "denied", "no waiver")),
        "full_waiver": any(t in norm for t in ("full waiver", "waived in full", "fully waived", "waiver granted in full")),
        "percent_waiver": None,
        "to_pay": [],
        "waived": [],
    }

    pct = re.search(r"(\d{1,3}(?:\.\d+)?)\s*%\s*waiver", text, flags=re.IGNORECASE)
    if pct:
        try:
            p = float(pct.group(1))
            if 0.0 <= p <= 100.0:
                out["percent_waiver"] = p
        except Exception:
            pass

    amt_re = re.compile(r"(?<!\d)(\d{1,3}(?:,\d{3})+|\d+(?:\.\d+)?)(?:\s*([kKmM]))?(?!\d)")
    for match in amt_re.finditer(text):
        amt = parse_amount((match.group(1) or "") + (match.group(2) or ""))
        if amt <= 0:
            continue
        after_text = text[match.end() : match.end() + 24].lower()
        if re.match(r"^\s*(day|days|week|weeks|month|months|year|years)\b", after_text):
            continue
        ctx_start = max(0, match.start() - 48)
        ctx_end = min(len(text), match.end() + 48)
        ctx = normalize(text[ctx_start:ctx_end])
        if "reinstatement" in ctx or ("filing" in ctx and "fee" in ctx):
            continue
        if "to pay" in ctx or "payable" in ctx or "pay " in ctx:
            out["to_pay"].append(amt)
        elif "waiv" in ctx or "grant" in ctx or "approve" in ctx:
            out["waived"].append(amt)
    return out


def compute_financials(comment: str, total_penalty: float, penalty_paid: float, requested_waiver: float) -> tuple[float, float, str]:
    total_penalty = max(0.0, total_penalty)
    penalty_paid = max(0.0, penalty_paid)
    requested_waiver = max(0.0, requested_waiver)
    remaining = max(0.0, total_penalty - penalty_paid)
    if remaining <= 0 and requested_waiver > 0:
        remaining = requested_waiver

    parsed = extract_comment_amounts(comment)
    granted = 0.0

    if remaining <= 0:
        granted = 0.0
    elif parsed["rejected"]:
        granted = 0.0
    elif parsed["full_waiver"]:
        granted = remaining
    elif isinstance(parsed["percent_waiver"], (int, float)):
        granted = remaining * (float(parsed["percent_waiver"]) / 100.0)
    elif parsed["waived"]:
        granted = max(parsed["waived"])
    elif parsed["to_pay"]:
        granted = max(0.0, remaining - min(parsed["to_pay"]))
    else:
        granted = min(requested_waiver, remaining)

    granted = max(0.0, min(granted, remaining))
    balance = max(0.0, remaining - granted)
    if balance <= 1e-9:
        status = "Fully Paid"
    elif penalty_paid <= 0 and granted <= 0:
        status = "No Payment"
    else:
        status = "Partially Paid"
    return granted, balance, status


def header_index_map(ws) -> dict[str, int]:
    out = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        label = normalize(val)
        if label:
            out[label] = c
    return out


def find_col(hmap: dict[str, int], terms: tuple[str, ...], exclude: tuple[str, ...] = ()) -> int | None:
    for label, idx in hmap.items():
        if all(t in label for t in terms) and all(x not in label for x in exclude):
            return idx
    return None


def drop_unnamed_header_columns(ws) -> int:
    to_drop = []
    for c in range(ws.max_column, 0, -1):
        val = str(ws.cell(row=1, column=c).value or "").strip().lower()
        if re.fullmatch(r"unnamed(?:[\s_:]*[0-9]+)?", val):
            to_drop.append(c)
    for c in to_drop:
        ws.delete_cols(c, 1)
    return len(to_drop)


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    backup_path = WORKBOOK_PATH.with_suffix(f".backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    copy2(WORKBOOK_PATH, backup_path)

    wb = load_workbook(WORKBOOK_PATH)
    total_rows_changed = 0
    total_cols_dropped = 0
    touched_sheets = 0

    for sheet in TARGET_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        touched_sheets += 1
        total_cols_dropped += drop_unnamed_header_columns(ws)
        hmap = header_index_map(ws)

        total_penalty_col = find_col(hmap, ("total", "penalty"))
        penalty_paid_col = find_col(hmap, ("penalty", "paid"))
        requested_waiver_col = find_col(hmap, ("requested", "waiver")) or find_col(hmap, ("request", "waiver"))
        balance_col = hmap.get("balance")
        if balance_col is None:
            balance_col = find_col(hmap, ("balance",), exclude=("pending",))
        if balance_col is None:
            balance_col = find_col(hmap, ("balance",))
        status_col = hmap.get("status") or find_col(hmap, ("status",), exclude=("file", "files"))
        committee_col = find_col(hmap, ("committee", "comment"))
        if committee_col is None:
            committee_col = find_col(hmap, ("comment",))

        granted_col = find_col(hmap, ("granted", "waiver"))
        if granted_col is None:
            granted_col = ws.max_column + 1
            ws.cell(row=1, column=granted_col).value = "Granted Waiver Amount"

        if balance_col is None:
            balance_col = ws.max_column + 1
            ws.cell(row=1, column=balance_col).value = "Balance"

        if status_col is None:
            status_col = ws.max_column + 1
            ws.cell(row=1, column=status_col).value = "status"

        for r in range(2, ws.max_row + 1):
            total_penalty = parse_amount(ws.cell(r, total_penalty_col).value) if total_penalty_col else 0.0
            penalty_paid = parse_amount(ws.cell(r, penalty_paid_col).value) if penalty_paid_col else 0.0
            requested_waiver = parse_amount(ws.cell(r, requested_waiver_col).value) if requested_waiver_col else 0.0
            comment = str(ws.cell(r, committee_col).value or "") if committee_col else ""

            if total_penalty <= 0 and penalty_paid <= 0 and requested_waiver <= 0 and not comment.strip():
                continue

            granted, balance, status = compute_financials(comment, total_penalty, penalty_paid, requested_waiver)

            before = (
                str(ws.cell(r, granted_col).value or "").strip(),
                str(ws.cell(r, balance_col).value or "").strip(),
                str(ws.cell(r, status_col).value or "").strip(),
            )
            ws.cell(r, granted_col).value = format_amount(granted)
            ws.cell(r, balance_col).value = format_amount(balance)
            ws.cell(r, status_col).value = status
            after = (
                str(ws.cell(r, granted_col).value or "").strip(),
                str(ws.cell(r, balance_col).value or "").strip(),
                str(ws.cell(r, status_col).value or "").strip(),
            )
            if after != before:
                total_rows_changed += 1

    wb.save(WORKBOOK_PATH)
    wb.close()
    print(f"Updated workbook: {WORKBOOK_PATH}")
    print(f"Backup created: {backup_path}")
    print(f"Sheets touched: {touched_sheets}")
    print(f"Unnamed columns dropped: {total_cols_dropped}")
    print(f"Rows changed: {total_rows_changed}")


if __name__ == "__main__":
    main()
